from copy import deepcopy
from itertools import combinations
from typing import Dict, List, Tuple, Literal

from openpyxl import Workbook

from utils import nlist_to_ntup, ntup_to_nlist
import mysql.connector

IntBool = Literal[0, 1]
"""
GameState:
A 2D array where the rows denote the symbols and the columns denote the size of the stacks of that symbol.
i.e. [[4, 1], [2, 2], [2, 1], []]. Represents the following board state:
A 4 stack and 1 stack of symbol A
two 2 stacks of symbol B
a 2 stack and 1 stack of symbol C
and no stack with symbol D

immutatable type Game State Tuple is needed for sets
Uses nlist_to_ntup and ntup_to_nlist to convert between the two types.
"""
GameState = List[List[int]]
GameStateTuple = Tuple[Tuple[int, ...]]
NUM_SYMBOLS = 4
NUM_TILES = 12
STARTING_CONFIGURATIONS: List[GameState] = [[[1, 1, 1], [1, 1, 1], [1, 1, 1], [1, 1, 1]],
    [[1, 1, 1, 1], [1, 1, 1], [1, 1, 1], [1, 1]],
    [[1, 1, 1, 1], [1, 1, 1, 1], [1, 1], [1, 1]],
    [[1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1], [1,]],
    [[1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1, 1], []],
    [[1, 1, 1, 1, 1], [1, 1, 1], [1, 1], [1, 1]],
    [[1, 1, 1, 1, 1], [1, 1, 1], [1, 1, 1], [1,]],
    [[1, 1, 1, 1, 1], [1, 1, 1, 1], [1, 1], [1,]],
    [[1, 1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1], []],
    [[1, 1, 1, 1, 1], [1, 1, 1, 1, 1], [1,], [1,]],
    [[1, 1, 1, 1, 1], [1, 1, 1, 1, 1], [1, 1], []],
    [[1, 1, 1, 1, 1, 1], [1, 1], [1, 1], [1, 1]],
    [[1, 1, 1, 1, 1, 1], [1, 1, 1], [1, 1], [1,]],
    [[1, 1, 1, 1, 1, 1], [1, 1, 1], [1, 1, 1], []],
    [[1, 1, 1, 1, 1, 1], [1, 1, 1, 1], [1,], [1,]],
    [[1, 1, 1, 1, 1, 1], [1, 1, 1, 1], [1, 1], []]
]

# MEMOIZATION: Dict[GameStateTuple, int] = {}
# Soluna Database Table Columns
ID = 0
STATE = 1
EVAL = 2
IS_DETERMINED = 3

config = {
    'user': 'root',
    'host': 'localhost',
    'database': 'soluna',
}
conn = None
cursor = None

def get_total_stacks(board: GameState) -> int:
    """
    Get the number of stacks in a board

    Parameters:
    - board: The game state

    Returns:
    - The total number of stacks in the board
    """
    return sum(len(symbol) for symbol in board)

def get_move_num(board: GameState) -> int:
    """
    Get the number of moves in a board

    Parameters:
    - board: The game state

    Returns:
    - The number of moves this current board state will make.
      (i.e. starting states start will return 1)
    """
    return NUM_TILES-get_total_stacks(board) + 1

def is_player1_turn(board: GameState) -> IntBool:
    """
    Determine if it is player 1's turn

    Parameters:
    - board: The game state

    Returns:
    - 1 if it is player 1's turn, 0 otherwise
    """
    return get_move_num(board) % 2

def get_wanted_score(board: GameState) -> int:
    """
    Get the wanted score of the game by the current player

    Parameters:
    - board: The game state

    Returns:
    1 if it is player 1's turn, -1 otherwise
    """
    return 2 * is_player1_turn(board) - 1

class Soluna:
    """
    A class representing the Soluna game.

    Attributes:
    - board (GameState): The current state of the Soluna board.

    Methods:
    - __init__(self, board: GameState): Initializes a Soluna game.
        Also validates the provided board and normalizes the position.
    - validate_board(self, board: GameState): Validates the provided board.
    - normalize_position(self): Normalizes the position of the board.
    - display_board(self): Displays the current state of the Soluna board.
    - get_moves(self) -> List[GameState]: Get all possible moves from a given state.

    Raises:
    - ValueError: If the provided board is invalid.

    Examples:
    >>> correct_board = Soluna([[5,], [1, 2], [2, 2], []])
    >>> correct_board.display_board()
    2A 2A
    2B 1B
    5C
    >>> too_many_symbols = Soluna([[4, 1], [2, 2], [2, 1], [], []])
    Traceback (most recent call last):
        ...
    ValueError: Invalid board: Board does not have exactly 4 symbols.

    >>> too_many_tiles = Soluna([[3, 1, 1], [2, 2], [2, 1], [1,]])
    Traceback (most recent call last):
        ...
    ValueError: Invalid board: Board does not have exactly 12 tiles.

    >>> non_positive_stack = Soluna([[4, 1], [2, 2], [2, 1], [0,]])
    Traceback (most recent call last):
        ...
    ValueError: Invalid board: Board's stacks must be positive integers.
    """
    def __init__(self, board: GameState) -> None:
        """
        Initialize a Soluna game.

        Parameters:
        - board: The Game State

        Raises:
        - ValueError: If the provided board is invalid.
        """
        self.validate_board(board)
        self.board = deepcopy(board)
        self.normalize_position()

    def validate_board(self, board: GameState) -> None:
        """
        Validate the provided board.

        Parameters:
        - board: The Game State to be verified

        Raises:
        - ValueError: If the provided board is invalid.

        The validation checks include:
        1. Ensuring the board has correct amount of symbols.
        2. Verifying that the board has correct amount of tiles.
        3. Confirming that each stack has positive height.
        """
        if len(board) != NUM_SYMBOLS:
            raise ValueError(f"Invalid board: Board does not have exactly {NUM_SYMBOLS} symbols.")

        if sum(sum(symbol) for symbol in board if symbol) != NUM_TILES:
            raise ValueError(f"Invalid board: Board does not have exactly {NUM_TILES} tiles.")

        if not all(isinstance(stack, int) and stack >= 1 for symbol in board for stack in symbol):
            raise ValueError("Invalid board: Board's stacks must be positive integers.")

    def display_board(self) -> None:
        """
        Display the current state of the Soluna board.
        Each stacks is seperated by spaces and consists of a number and letter.
        The number is the size of the stack and the letter is the the top stack's symbol
        """
        labels = iter("ABCD")
        for symbol in self.board:
            label = next(labels)
            if symbol:
                print(" ".join(f"{stack}{label}" for stack in symbol))

    def normalize_position(self) -> None:
        """
        Transform self.board represented by a 2d array into a normalized equivalent 2d array.
        1. Each symbols stack sizes are given in a nonincreasing order.
        2. The number of stacks the symbols have must be in a nonincreasing order.
        3. Two symbols with the same amount of stacks must be in reverse lexicographical ordering
        ex. ((1, 1, 1), (3, 1), (2, 2), (1,))
        """
        for i in range(NUM_SYMBOLS):
            self.board[i] = sorted(self.board[i], reverse=True)

        self.board = sorted(self.board, key=lambda symbol: (len(symbol), symbol), reverse=True)

    def get_moves(self) -> List[GameState]:
        """
        Get all possible moves from a given state.
        """
        possible_moves: List[GameState] = []
        board_copy: GameState = deepcopy(self.board)

        # combining stacks of same symbol
        for index, symbol in enumerate(self.board):
            distinct = set(combinations(symbol, 2))
            for (stack1, stack2) in distinct:
                self.board[index].remove(stack1)
                self.board[index].remove(stack2)
                self.board[index].append(stack1+stack2)
                self.normalize_position()
                possible_moves.append(self.board)
                self.board = deepcopy(board_copy)

        # combining stacks of different symbol
        combinations_2 = list(combinations(range(NUM_SYMBOLS), 2))
        for (symbol1, symbol2) in combinations_2:
            matching_numbers = set(self.board[symbol1]) & set(self.board[symbol2])
            for num in matching_numbers:
                self.board[symbol1].remove(num)
                self.board[symbol2].remove(num)
                self.board[symbol1].append(num*2)
                self.normalize_position()
                possible_moves.append(self.board)
                self.board = deepcopy(board_copy)

                self.board[symbol1].remove(num)
                self.board[symbol2].remove(num)
                self.board[symbol2].append(num*2)
                self.normalize_position()
                possible_moves.append(self.board)
                self.board = deepcopy(board_copy)

        unique_moves = []
        for move in possible_moves:
            if move not in unique_moves:
                unique_moves.append(move)
        return unique_moves

def get_all_positions() -> List[GameState]:
    """
    Use breadth-first search to find all possible game states by move.
    """
    possible_positions_by_move: list[set[GameStateTuple]] = []
    # move 1 possible positions
    possible_positions_by_move.append(set(nlist_to_ntup(config) for config in STARTING_CONFIGURATIONS))

    # move 2-12 possible positions
    for positions_by_move in possible_positions_by_move:
        new_possible_positions = set()
        for position in positions_by_move:
            soluna_game = Soluna(ntup_to_nlist(position))
            new_possible_positions.update(nlist_to_ntup(soluna_game.get_moves()))
        if len(new_possible_positions) != 0:
            possible_positions_by_move.append(new_possible_positions)

    # convert list of set into flattened list in reverse move order
    return [ntup_to_nlist(position) for positions in possible_positions_by_move for position in positions]

def solve(board: GameState) -> int:
    """
    Solve a position using memoization.
    """
    soluna_game = Soluna(board)

    cursor.execute(f'SELECT * FROM soluna WHERE state = "{soluna_game.board}"')
    board_data = cursor.fetchone()

    if board_data:
        return board_data[EVAL]

    possible_moves = soluna_game.get_moves()
    wanted_score = get_wanted_score(board)

    # if len(possible_moves) == 0:
    #     MEMOIZATION[board_key] = -2*wanted_score
    #     print(board,"No moves left")
    #     return MEMOIZATION[board_key]

    possibilities = [solve(move) for move in possible_moves]
    # if set(possibilities) == {2} or set(possibilities) == {-2}:
    #     print("Guranteed outcome", board)
    #     MEMOIZATION[board_key] = possibilities[0]
    #     return possibilities[0]

    # get probability of player 1 winning by taking the percentage of positive in prossiilities
    # player1_wins = sum(1 for possibility in possibilities if possibility > 0)
    # probability_player1_wins = player1_wins/len(possibilities)
    # print(board, probability_player1_wins, possibilities)


    if wanted_score in possibilities:
        cursor.execute(f'INSERT INTO soluna (state, eval) VALUES ("{soluna_game.board}", {wanted_score})')
        conn.commit()
        return wanted_score

    cursor.execute(f'INSERT INTO soluna (state, eval) VALUES ("{soluna_game.board}", {-wanted_score})')
    conn.commit()
    return -wanted_score


def solve_game() -> None:
    """
    Solve the game using memoization

    Where
    first player victory = 1
    second player victory = -1
    """
    all_positions = get_all_positions()

    for position in all_positions[::-1]:
        print(f"Position {len(all_positions)-all_positions.index(position)}/{len(all_positions)}")
        solve(position)


# def make_sheet(file_name: str) -> None:
#     """
#     Make a sheet
#     """
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Moves"

#     sheet.cell(row=1, column=1, value='Move')
#     sheet.cell(row=1, column=2, value='Nested Tuple')
#     sheet.cell(row=1, column=3, value='Value')

#     next_row = 2
#     for board, value in MEMOIZATION.items():
#         sheet.cell(row=next_row, column=1, value=get_move_num(ntup_to_nlist(board)))
#         sheet.cell(row=next_row, column=2, value=str(board))
#         sheet.cell(row=next_row, column=3, value=value)
#         next_row += 1

#     sheet.column_dimensions['B'].width = 30

#     workbook.save(f'{file_name}.xlsx')

def main() -> None:
    """
    Main function
    """
    global conn, cursor
    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()
    except mysql.connector.Error as e:
        print(f"Error connecting to MySQL: {e}")
        return
    solve_game()
    # make_sheet("memoization")
    if 'conn' in locals() and conn.is_connected():
        cursor.close()
        conn.close()
        print('MySQL connection closed')

if __name__ == '__main__':
    main()
