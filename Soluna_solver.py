from typing import Dict, List, Tuple
from copy import deepcopy
from itertools import combinations
from openpyxl import Workbook

class Soluna:
    """
    A class representing the Soluna game.

    GameState:
    A 2D array where the rows denote the symbols and the columns denote the size of the stacks of that symbol.
    i.e. [[4, 1], [2, 2], [2, 1], []]. Represents the following board state:
    A 4 stack and 1 stack of symbol A
    two 2 stacks of symbol B
    a 2 stack and 1 stack of symbol C
    and no stack with symbol D

    Attributes:
    - board (GameState): The current state of the Soluna board.

    Methods:
    - __init__(self, board: GameState): Initializes a Soluna game.
    - validate_board(self, board: GameState): Validates the provided board.
    - display_board(self): Displays the current state of the Soluna board.

    Raises:
    - ValueError: If the provided board is invalid.

    Examples:
    >>> correct_board = Soluna([[5,], [1, 2], [2, 2], []])
    >>> correct_board.display_board()
    2A 2A
    2B 1B
    5C
    >>> too_many_symbols = Soluna([[4, 1], [2, 2], [2, 1], [], []])
    Traceback (most recent call last):
        ...
    ValueError: Invalid board: Board does not have exactly 4 symbols.

    >>> too_many_tiles = Soluna([[3, 1, 1], [2, 2], [2, 1], [1,]])
    Traceback (most recent call last):
        ...
    ValueError: Invalid board: Board does not have exactly 12 tiles.

    >>> non_positive_stack = Soluna([[4, 1], [2, 2], [2, 1], [0,]])
    Traceback (most recent call last):
        ...
    ValueError: Invalid board: Board's stacks must be positive integers.
    """
    GameStateTuple = Tuple[Tuple[int]]
    GameState = List[List[int]]
    NUM_SYMBOLS = 4
    NUM_TILES = 12
    STARTING_CONFIGURATIONS = [[[1, 1, 1], [1, 1, 1], [1, 1, 1], [1, 1, 1]],
        [[1, 1, 1, 1], [1, 1, 1], [1, 1, 1], [1, 1]],
        [[1, 1, 1, 1], [1, 1, 1, 1], [1, 1], [1, 1]],
        [[1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1], [1,]],
        [[1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1, 1], []],
        [[1, 1, 1, 1, 1], [1, 1, 1], [1, 1], [1, 1]],
        [[1, 1, 1, 1, 1], [1, 1, 1], [1, 1, 1], [1,]],
        [[1, 1, 1, 1, 1], [1, 1, 1, 1], [1, 1], [1,]],
        [[1, 1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1], []],
        [[1, 1, 1, 1, 1], [1, 1, 1, 1, 1], [1,], [1,]],
        [[1, 1, 1, 1, 1], [1, 1, 1, 1, 1], [1, 1], []],
        [[1, 1, 1, 1, 1, 1], [1, 1], [1, 1], [1, 1]],
        [[1, 1, 1, 1, 1, 1], [1, 1, 1], [1, 1], [1,]],
        [[1, 1, 1, 1, 1, 1], [1, 1, 1], [1, 1, 1], []],
        [[1, 1, 1, 1, 1, 1], [1, 1, 1, 1], [1,], [1,]],
        [[1, 1, 1, 1, 1, 1], [1, 1, 1, 1], [1, 1], []]
    ]
    minimax_counter = 0
    memoization_dict_by_move: "Dict[GameStateTuple: (int, GameState)]" = [{} for _ in range(12)]


    @classmethod
    def update_memoization_dict(cls, move_moved: int, board_key: GameState, value: int) -> None:
        """
        Updates the memoization dictionary for a specific move and board key.

        Parameters:
        - move_moved: The move number.
        - board_key: The key representing the current board state.
        - value: The value associated with the board state.

        """
        cls.memoization_dict_by_move[move_moved][board_key] = value

    def __init__(self, board: GameState) -> None:
        """
        Initialize a Soluna game.

        Parameters:
        - board: The Game State

        Raises:
        - ValueError: If the provided board is invalid.
        """
        self.validate_board(board)
        self.board = board
        self.normalize_position()

    def validate_board(self, board: GameState) -> None:
        """
        Validate the provided board.

        Parameters:
        - board: The Game State to be verified

        Raises:
        - ValueError: If the provided board is invalid.

        The validation checks include:
        1. Ensuring the board has correct amount of symbols.
        2. Verifying that the board has correct amount of tiles.
        3. Confirming that each stack has positive height.
        """
        if len(board) != Soluna.NUM_SYMBOLS:
            raise ValueError(f"Invalid board: Board does not have exactly {Soluna.NUM_SYMBOLS} symbols.")

        if sum(sum(symbol) for symbol in board if symbol) != Soluna.NUM_TILES:
            raise ValueError(f"Invalid board: Board does not have exactly {Soluna.NUM_TILES} tiles.")

        if not all(isinstance(stack, int) and stack >= 1 for symbol in board for stack in symbol):
            raise ValueError("Invalid board: Board's stacks must be positive integers.")

    def display_board(self) -> None:
        """
        Display the current state of the Soluna board.
        Each stacks is seperated by spaces and consists of a number and letter.
        The number is the size of the stack and the letter is the the top stack's symbol
        """
        labels = iter("ABCD")
        for symbol in self.board:
            label = next(labels)
            if symbol:
                print(" ".join(f"{stack}{label}" for stack in symbol))

    def normalize_position(self) -> None:
        """
        Transform self.board represented by a 2d array into a normalized equivalent 2d array.
        1. Each symbols stack sizes are given in a nonincreasing order.
        2. The number of stacks the symbols have must be in a nonincreasing order.
        3. Two symbols with the same amount of stacks must be in reverse lexicographical ordering
        ex. ((1, 1, 1), (3, 1), (2, 2), (1,))
        """
        for i in range(Soluna.NUM_SYMBOLS):
            self.board[i] = sorted(self.board[i], reverse=True)

        self.board = sorted(self.board, key=lambda symbol: (len(symbol), symbol), reverse=True)

    def get_moves(self) -> List[GameState]:
        """
        Get all possible moves from a given state.
        """
        possible_moves = []
        board_copy = deepcopy(self.board)

        # combining stacks of same symbol
        for index, symbol in enumerate(self.board):
            distinct = set(combinations(symbol, 2))
            for (stack1, stack2) in distinct:
                self.board[index].remove(stack1)
                self.board[index].remove(stack2)
                self.board[index].append(stack1+stack2)
                self.normalize_position()
                possible_moves.append(self.board)
                self.board = deepcopy(board_copy)

        # combining stacks of different symbol
        combinations_2 = list(combinations(range(Soluna.NUM_SYMBOLS), 2))
        for (symbol1, symbol2) in combinations_2:
            matching_numbers = set(self.board[symbol1]) & set(self.board[symbol2])
            for num in matching_numbers:
                self.board[symbol1].remove(num)
                self.board[symbol2].remove(num)
                self.board[symbol1].append(num*2)
                self.normalize_position()
                possible_moves.append(self.board)
                self.board = deepcopy(board_copy)

                self.board[symbol1].remove(num)
                self.board[symbol2].remove(num)
                self.board[symbol2].append(num*2)
                self.normalize_position()
                possible_moves.append(self.board)
                self.board = deepcopy(board_copy)

        unique_moves = []

        for move in possible_moves:
            if move not in unique_moves:
                unique_moves.append(move)

        return unique_moves

    def minimax(self) -> int:
        """
        Implement the minimax algorithm with memoization to find the optimal move.
        Player 1 is the maximizing player (starts the game and wants the final position to have an odd number of stacks).
        Player 2 is the minimizing player (goes second and wants the final position to have an even number of stacks).

        Returns:
        - The minimax value for the current state of the board.
        """
        Soluna.minimax_counter += 1

        move_moved = Soluna.NUM_TILES-sum(len(symbol) for symbol in self.board if symbol)
        is_player1_turn  = 1 - move_moved % 2
        possible_moves = self.get_moves()

        board_key: Soluna.GameStateTuple = list_to_tuple(self.board)
        if board_key in Soluna.memoization_dict_by_move[move_moved]:
            return Soluna.memoization_dict_by_move[move_moved][board_key][0]

        # Base cases
        if len(possible_moves) == 0:
            self.update_memoization_dict(move_moved, board_key, -2 * is_player1_turn + 1)
            return -2*is_player1_turn+1

        board_copy = deepcopy(self.board)
        if is_player1_turn:
            max_eval = float('-inf')
            for move in possible_moves:
                self.board = move
                eval = self.minimax()
                self.board = deepcopy(board_copy)
                max_eval = max(max_eval, eval)
            board_key = list_to_tuple(self.board)
            self.update_memoization_dict(move_moved, board_key, (max_eval, None))
            return max_eval
        else:
            min_eval = float('inf')
            for move in possible_moves:
                self.board = move
                eval = self.minimax()
                self.board = deepcopy(board_copy)
                min_eval = min(min_eval, eval)
            self.update_memoization_dict(move_moved, board_key, (min_eval, None))
            return min_eval

    def solve_game() -> None:
        """
        Update the memoization dictionary to hold the winner under optimal play for every position.
        """
        for board in Soluna.STARTING_CONFIGURATIONS:
            board_copy = deepcopy(board)
            soluna_game = Soluna(board_copy)
            soluna_game.display_board()
            soluna_game.minimax()
            print()

    def get_all_positions() -> List[List[GameState]]:
        """
        Use breadth-first search to find all possible game states by move.
        """
        possible_positions_by_move = []
        possible_positions_by_move.append(set(list_to_tuple(config) for config in Soluna.STARTING_CONFIGURATIONS))

        for positions_by_move in possible_positions_by_move:
            new_possible_positions = set()
            for position in positions_by_move:
                soluna_game = Soluna(tuple_to_list(position))
                new_possible_positions.update(list_to_tuple(soluna_game.get_moves()))
            if len(new_possible_positions) != 0:
                possible_positions_by_move.append(new_possible_positions)
        return possible_positions_by_move

    def make_sheet(file_name: str) -> None:
        """
        Make a sheet
        """
        workbook = Workbook()

        for move, memo_dict in enumerate(Soluna.memoization_dict_by_move):
            sheet = workbook.create_sheet(title=f'Move_{move+1}')

            sheet.cell(row=1, column=1, value='Nested Tuple')
            sheet.cell(row=1, column=2, value='Value')

            for row, (gameState, value) in enumerate(memo_dict.items(), start=2):
                sheet.cell(row=row, column=1, value=str(gameState))
                sheet.cell(row=row, column=2, value=value)

            sheet.column_dimensions['A'].width = 30

        workbook.remove(workbook['Sheet'])
        workbook.save(f'{file_name}.xlsx')


def list_to_tuple(input_list):
    """
    Recursively convert a nested list to a nested tuple.
    """
    if isinstance(input_list, list):
        return tuple(list_to_tuple(item) for item in input_list)
    else:
        return input_list

def tuple_to_list(input_tuple):
    """
    Recursively convert a nested tuple to a nested list.
    """
    if isinstance(input_tuple, tuple):
        return [tuple_to_list(item) for item in input_tuple]
    else:
        return input_tuple

# example_board = [
#     [1, 1, 1, 1],
#     [1, 1, 1, 1],
#     [2, 1, 1],
#     []
# ]


Soluna.solve_game()
Soluna.make_sheet("wow!")

import doctest
doctest.testmod()
